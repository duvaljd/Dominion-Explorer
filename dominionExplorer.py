#  IMPORT STATEMENTS #
from bs4 import BeautifulSoup
from creds import *
from datetime import datetime
import json
import html5lib
import html2text
textMaker = html2text.HTML2Text()
textMaker.images_to_alt = True
textMaker.unicode_snob = True
textMaker.ignore_emphasis = True
textMaker.body_width = 0
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import psycopg2
import psycopg2.extras
import random
import requests
import sys, os
import unittest
#  /IMPORT STATEMENTS #

#  GLOBAL VARIABLES #
BASEURL ="http://wiki.dominionstrategy.com"

CACHE_FNAME = 'dominion_cache.json'

DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S.%f"

DB_CONNECTION = None
DB_CURSOR = None

DEBUG = True
GETNEWDATA = True

promptMsg = ""
originalPrompt = "\nPlease the type the name of a Dominion card (case-sensitive) or set you would like to know about, type 'help' for a list of other commands, or 'exit' to leave the program.\n"

DONE = False
localExit = True
resetPrompt = True
# /GLOBAL VARIABLES #

#  DATABASE SETUP #
##  Connection & Cursor ##
def getConnection_andCursor():
	global DB_CONNECTION, DB_CURSOR
	if not DB_CONNECTION:
		try:
			DB_CONNECTION = psycopg2.connect("dbname='{}' user='{}' password='{}'".format(DB_NAME, DB_USER, DB_PASS))
			print("Connected to database {} as {}.".format(DB_NAME, DB_USER))

		except:
			print("Unable to connect to database {}.".format(DB_NAME))
			sys.exit(1)

	if not DB_CURSOR:
		DB_CURSOR = DB_CONNECTION.cursor(cursor_factory = psycopg2.extras.RealDictCursor)

	return DB_CONNECTION, DB_CURSOR
## /Connection & Cursor ##

##  Create Tables ##

def makeTables():

	# Clear tables # 
	if DEBUG:
		print("\n*************** Dropping tables ***************")
	else:
		print("\nBuilding database...", flush=True)

	try:
		cur.execute(""" DROP TABLE IF EXISTS cards CASCADE """)
		conn.commit()

		if DEBUG:
			print("     - 'cards' table dropped.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could not drop 'cards' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))

	try:
		cur.execute(""" DROP TABLE IF EXISTS csets CASCADE """)
		conn.commit()

		if DEBUG:
			print("     - 'csets' table dropped.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could not drop 'csets' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))

	try:
		cur.execute(""" DROP TABLE IF EXISTS recommendations CASCADE """)
		conn.commit()
		if DEBUG:
			print("     - 'recommendations' table dropped.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could drop 'recommendations' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))

	try:
		cur.execute(""" DROP TABLE IF EXISTS cardsInRecs CASCADE """)
		conn.commit()

		if DEBUG:
			print("     - 'cardsInRecs' table dropped.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could drop 'cardsInRecs' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))
	if DEBUG:
		print("***********************************************\n")

	# Make tables #
		print("\n*************** Adding tables ***************")

	try:
		cur.execute(""" CREATE TABLE IF NOT EXISTS csets(

			sid SERIAL,
			name VARCHAR(40) NOT NULL,
			cardnumber VARCHAR(100),
			themes VARCHAR(100),
			release VARCHAR(100),
			coverart VARCHAR(100),

			PRIMARY KEY (sid)

			)
		""")
		conn.commit()

		if DEBUG:
			print("     - 'csets' table created successfully.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could not create 'csets' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))

	try:
		cur.execute(""" CREATE TABLE IF NOT EXISTS cards(

			cid SERIAL NOT NULL,
			sid INT NOT NULL,  
			name VARCHAR(100) NOT NULL UNIQUE,
			cost VARCHAR (100),
			types VARCHAR (100),
			illustrators VARCHAR (100),
			description TEXT,

			PRIMARY KEY (cid),

			FOREIGN KEY (sid) REFERENCES csets(sid)

			)
		""")
		conn.commit()

		if DEBUG:
			print("     - 'cards' table created successfully.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could not create 'cards' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))

	try:
		cur.execute(""" CREATE TABLE IF NOT EXISTS recommendations(
			rid SERIAL,
			name VARCHAR(100) NOT NULL UNIQUE,
			set1 INTEGER NOT NULL,
			set2 INTEGER,
			set3 INTEGER,

			PRIMARY KEY (rid),
			FOREIGN KEY (set1) REFERENCES csets(sid),
			FOREIGN KEY (set2) REFERENCES csets(sid),
			FOREIGN KEY (set3) REFERENCES csets(sid)

			)
		""")
		conn.commit()

		if DEBUG:
			print("     - 'recommendations' table created successfully.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could not create 'recommendations' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))

	try:
		cur.execute(""" CREATE TABLE IF NOT EXISTS cardsInRecs(
			cid INTEGER NOT NULL,
			rid INTEGER NOT NULL,

			FOREIGN KEY (cid) REFERENCES cards(cid),
			FOREIGN KEY (rid) REFERENCES recommendations(rid),
			UNIQUE(cid, rid)

			)
		""")
		conn.commit()

		if DEBUG:
			print("     - 'cardsInRecs' table created successfully.")

	except Exception as e:
		if DEBUG:
			exc_type, exc_obj, exc_tb = sys.exc_info()
			fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
			print("     * Could not create 'cardsInRecs' table.")
			print("({}{}{})".format(exc_type, fname, exc_tb.tb_lineno))
	if DEBUG:
		print("***********************************************\n")
	else:
		print("...Done!")
## /Create Tables ##
#  /DATABASE SETUP #

#  CACHE SETUP #
##  Cache Files ##
try:
	with open(CACHE_FNAME, "r") as cache_file:
		cache_json = cache_file.read()
		CACHE_DICTION = json.loads(cache_json)
except:
	CACHE_DICTION = {}
## /Cache Files ##

##  Cache Functions ##
def has_cache_expired(timestamp_str, expire_in_days):
    now = datetime.now()
    cache_timestamp = datetime.strptime(timestamp_str, DATETIME_FORMAT)
    delta = now - cache_timestamp
    delta_in_days = delta.days

    if delta_in_days > expire_in_days:
        return True
    else:
        return False

def get_from_cache(cacheKey, dictionary):
    cacheKey = cacheKey.lower()
    if cacheKey in dictionary:
        data_assoc_dict = dictionary[cacheKey]
        if has_cache_expired(data_assoc_dict['timestamp'],data_assoc_dict["expire_in_days"]):
            if DEBUG:
                print("Cache has expired for {}".format(cacheKey))
            del dictionary[cacheKey]
            data = None
        else:
            data = dictionary[cacheKey]['values']
    else:
        data = None
    return(data)

def set_in_data_cache(cacheKey, url, data, expire_in_days):
	cacheKey = cacheKey.lower()
	CACHE_DICTION[cacheKey] = {
		'url': url,
		'values': data,
		'timestamp': datetime.now().strftime(DATETIME_FORMAT),
		'expire_in_days': expire_in_days
	}

	with open(CACHE_FNAME, 'w') as cache_file:
		cache_json = json.dumps(CACHE_DICTION)
		cache_file.write(cache_json)

	if DEBUG:
		print("Saved data from {} to {}.".format(url, CACHE_FNAME))

## /Cache Functions ##
# /CACHE SETUP #

#  FUNCTIONS #
##  Data Collection ##
###  get html ###
def getData(url, expire_in_days = 31, dataType = None):
	data = get_from_cache("{} {}".format(url, dataType), CACHE_DICTION)

	if data:
		if DEBUG:
			print("Loaded ({}) data from cache.".format(url), flush=True)

	else:
		if DEBUG:
			time = requests.get(url).elapsed.total_seconds()
			print("Fetched new data from ({}) in {}.".format(url, time), flush=True)

		resp = requests.get(url)
		htmlContent = resp.text.encode('utf-8')
		content = ""
		cacheKey = ""

		try:
			if dataType == "setList":
				soup = BeautifulSoup(htmlContent, "html5lib")
				content = soup.find('span', id="Releases").findNext('ul')

			elif dataType == "set":
				soup = BeautifulSoup(htmlContent, "html5lib")
				content = soup.find('div', id="mw-content-text").findNext('table')

			elif dataType == "card":
				soup = BeautifulSoup(htmlContent, "html5lib")
				content = soup.find('div', id="mw-content-text").findNext('table')

			elif dataType == "rec":
				soup = BeautifulSoup(htmlContent, "html5lib")
				for tag in soup.find('span', attrs={"class":"mw-headline", "id":"Recommended_Sets_of_10"}).find_parent('h2').next_siblings:

					if tag.name == "h2":
						break
					elif tag.name == "p" and tag.findChild() is not None:
						break
					else:
						content += str(tag)

			cacheKey = "{} {}".format(url,dataType)
			data = json.dumps(str(content))
			set_in_data_cache(cacheKey, url, data, expire_in_days)

		except Exception as e:
			if DEBUG:
				print(e)

	return(json.loads(data))
### /get html ###

###  create link lists ###
def getSetLinks():
	data = getData("http://wiki.dominionstrategy.com/index.php/Sets", dataType = "setList")
	soup = BeautifulSoup(data, "html5lib")
	setLinks = []

	for li in soup.find_all("li"):
		for link in li.find_all("a"):
			if link["href"] not in setLinks:
				setLinks.append(link["href"])

	return(setLinks)

def getCardLinks():
	data = getData("http://wiki.dominionstrategy.com/index.php/List_of_cards", dataType = "card")
	soup = BeautifulSoup(data, "html5lib")
	ignore = ['Hex', 'Event', 'Boon', 'Landmark', 'State', 'Promo', 'Artifact', 'Project']
	cardLinks = []

	for span in soup.find_all("span", attrs={"class":["card-popup"]}):
		for link in span.find_all("a"):
			if (link.findNext('td').text.strip() not in ignore and
				link.findNext('td').findNext('td').text.strip() not in ignore):
					if link["href"] not in cardLinks:
						cardLinks.append(link["href"])

	return(cardLinks)
### /create link lists ###

###  get data from links ###
def getSetData(listOf_setLinks):
	soupList = []
	
	for link in listOf_setLinks:

		try:
			data = getData("{}{}".format(BASEURL, link), dataType = "set")
		except:
			if DEBUG:
				print("Could not get data for set {}.".format(link[11:]), flush=True)

		soup = BeautifulSoup(data, "html5lib")
		soupList.append(soup)

	return(soupList)

def getCardData(listOf_cardLinks):
	soupList = []
	
	for link in listOf_cardLinks:

		try:
			data = getData("{}{}".format(BASEURL, link), dataType = "card")
		except Exception as e:
			if DEBUG:
				print("Could not get data for card {}.\n{}".format(link[11:], e), flush=True)

		soup = BeautifulSoup(data, "html5lib")
		soupList.append(soup)

	return(soupList)

def getRecData(listOf_setLinks):
	recsList = []

	for link in listOf_setLinks:
		try:
			data = getData("{}{}".format(BASEURL, link), dataType = "rec")
		except Exception as e:
			if DEBUG:
				print("Could not get recommendations for set {}.\n{}".format(link[11:], e), flush=True)

		soup = BeautifulSoup(data, "html5lib")
		recsList.append(soup)

	return(recsList)
### /get data from links ###

## /Data Collection ##

##  Data Insertion ##
###  insert a set ###
def insertSet(setSoup):
	name = setSoup.select('strong')[0].text
	cardnumber = setSoup.find('th', text='Cards').findNext('td').text
	themes = setSoup.find('th', text='Theme(s)').findNext('td').text
	release = setSoup.find('th', text='Release').findNext('td').text
	coverart = setSoup.find('th', text='Cover artist').findNext('td').text

	try:
		sql = """ INSERT INTO csets(name, cardnumber, themes, release, coverart) VALUES (%s, %s, %s, %s, %s) """
		cur.execute(sql, (name, cardnumber, themes, release, coverart))
		conn.commit()

		if DEBUG:
			print("Successfully inserted {} set into 'csets' table.".format(name))

	except:
		if DEBUG:
			print("Could not insert {} into 'csets' table.".format(name))
### /insert a set ###

###  insert a card ###
def insertCard(cardSoup):
	ignoredCards = ['Sauna', 'Avanto']
	name = cardSoup.find('strong').text

	if name not in ignoredCards:
		cost = cardSoup.find('a', title='Cost').findNext('img')['alt']
		types = cardSoup.find('a', title='Card types').findNext('td').text
		cset = cardSoup.find('a', title='Expansions').findNext('a').text
		if cset == "Base":
			cset = "Dominion"
		illustrators = cardSoup.find('th', text='Illustrator(s)').findNext('td').text
		dscSoup = cardSoup.find('th', text='Card text').findNext('td').prettify()
		dscText = textMaker.handle(dscSoup)
		dscCleanUp_imgs = dscText.replace('[ P ](/index.php/Potion "Potion")', 'Potion').replace('[ VP.png ](/index.php/Victory_point "Victory point")', 'Victory Point')
		dscCleanup_pxn = dscCleanUp_imgs.replace('_', '').replace(' $', '$').replace('|  ', '').replace('| ', '').replace('\\', '').replace('+ ', '+').replace('  .', '.').replace(' .', '.').replace('  ;', ';').replace(' ;', ';').replace(';  ', '; ').replace(':  ', ': ').replace(',  ', ', ').replace('  ', ' ')
		description = dscCleanup_pxn.strip()

		try:
			sql = """ INSERT INTO cards(name, sid, cost, types, illustrators, description) VALUES (%s, (SELECT sid FROM csets WHERE name = %s), %s, %s, %s, %s) """
			cur.execute(sql, (name, cset, cost, types, illustrators, description))
			conn.commit()

			if DEBUG:
				print("Successfully inserted {} into 'cards' table.".format(name))

		except Exception as e:
			if DEBUG:
				print("Could not insert {} into 'cards' table. Error: {}".format(name, e))

### /insert a card ###

###  insert a recommendation & card-rec relationships ###
def insertRec(recSoup):
	recSets= []

	knights = ['Dame Anna', 'Dame Josephine', 'Dame Molly', 'Dame Natalie', 'Dame Sylvia', 'Sir Bailey', 'Sir Destry', 'Sir Martin', 'Sir Michael', 'Sir Vander']
	castles = ['Humble Castle', 'Crumbling Castle', 'Small Castle', 'Haunted Castle', 'Opulent Castle', 'Sprawling Castle', 'Grand Castle', "King's Castle"]
	allSets = ['Dominion', 'Intrigue', 'Seaside', 'Alchemy', 'Prosperity', 'Cornucopia', 'Hinterlands', 'Dark', 'Guilds', 'Adventures', 'Empires', 'Nocturne', 'Renaissance']

	recCards = []
	recName = None

	for tag in recSoup.find('h3').next_siblings:
		if tag.name == "table":
			if tag.previous_sibling.previous_sibling.name == "h3":
				recSets = []
				for word in str(tag.previous_sibling.previous_sibling.contents[2].text).replace('/', ' ').split():
					if word in allSets:
						if word == "Dark":
							word = "Dark Ages"
						recSets.append(word)

			recCards = []
			recName = tag.select('tbody > tr')[0].select('th')[0].text.replace('[+/-]', '').strip()

			for link in tag.select('tbody > tr')[1].select('a'):
				if link.text == 'Knights':
					for knight in knights:
						recCards.append(knight)
				elif link.text == 'Castles':
					for castle in castles:
						recCards.append(castle)
				elif link.text == "ill-Gotten Gains":
					recCards.append("Ill-Gotten Gains")
				elif link.text == "Jack of all Trades":
					recCards.append("Jack of All Trades")
				else:
					recCards.append(link.text)

			for link in tag.select('tbody > tr')[4].select('a'):
				if link.text == 'Knights':
					for knight in knights:
						recCards.append(knight)
				elif link.text == 'Castles':
					for castle in castles:
						recCards.append(castle)
				elif link.text == "ill-Gotten Gains":
					recCards.append("Ill-Gotten Gains")
				elif link.text == "Jack of all Trades":
					recCards.append("Jack of All Trades")
				else:
					recCards.append(link.text)
			
			if len(recSets) == 1:
				try:
					sql = """ INSERT INTO recommendations(name, set1) VALUES (%s, (SELECT sid FROM csets WHERE name = %s)) ON CONFLICT DO NOTHING """
					cur.execute(sql, (recName, recSets[0]))
					conn.commit()
					if DEBUG:
						print("Set 1: {}".format(recSets[0]))
						# print("Successfully created {} and inserted first set reference.".format(recName))

				except Exception as e:
					if DEBUG:
						print("Error 1: {}".format(e))

			elif len(recSets) == 2:
				try:
					sql = """ INSERT INTO recommendations(name, set1, set2) VALUES (%s, (SELECT sid FROM csets WHERE name = %s), (SELECT sid FROM csets WHERE name = %s)) ON CONFLICT DO NOTHING """
					cur.execute(sql, (recName, recSets[0], recSets[1]))
					conn.commit()
				
					if DEBUG:
						print("Set 1 & 2: {}, {}".format(recSets[0], recSets[1]))

				except Exception as e:
					if DEBUG:
						print("Error 2: {}".format(e))

			elif len(recSets) == 3:
				try:
					sql = """ INSERT INTO recommendations(name, set1, set2, set3) VALUES (%s, (SELECT sid FROM csets WHERE name = %s), (SELECT sid FROM csets WHERE name = %s), (SELECT sid FROM csets WHERE name = %s)) ON CONFLICT DO NOTHING """
					cur.execute(sql, (recName, recSets[0], recSets[1], recSets[2]))
					conn.commit()
					
					if DEBUG:
						print("Set 1, 2, 3: {}, {}, {}".format(recSets[0]), recSets[1], recSets[3])

				except Exception as e:
					if DEBUG:
						print("Error 3: {}".format(e))

			for card in recCards:
				try:
					sql = """ INSERT INTO cardsInRecs(cid, rid) VALUES ((SELECT cid FROM cards WHERE name = %s),(SELECT rid FROM recommendations WHERE name = %s)) ON CONFLICT DO NOTHING """
					cur.execute(sql, (card, recName))
					conn.commit()

					if DEBUG:
						print("Successfully inserted card-set reference for {}-{}".format(card, recName))
				except Exception as e:
					if DEBUG:
						print("Could not insert card-set reference for {}-{}: {}".format(card, recName, e))
### /insert a recommendation & card-rec relationships ###
## /Data Insertion ##

##  Data Fetching ##
def fetchAll_cardIDs():
	allIDs = []

	sql = """ SELECT cid FROM cards WHERE cid IS NOT NULL """
	cur.execute(sql)
	results = cur.fetchall()

	for result in results:
		allIDs.append(result['cid'])

	return(allIDs)

def fetchAll_setIDs():
	allIDs = []

	sql = """ SELECT sid FROM csets WHERE sid IS NOT NULL """
	cur.execute(sql)
	results = cur.fetchall()

	for result in results:
		allIDs.append(result['sid'])

	return(allIDs)

def fetchAll_recIDs():
	allIDs = []

	sql = """ SELECT rid FROM recommendations WHERE rid IS NOT NULL """
	cur.execute(sql)
	results = cur.fetchall()

	for result in results:
		allIDs.append(result['rid'])

	return(allIDs)

def makeList():
	cardObjs = []
	setObjs = []
	recObjs = []

	for cid in allCards:
		cardObjs.append(Card(cid))

	for sid in allSets:
		setObjs.append(Set(sid))

	for rid in allRecs:
		recObjs.append(Rec(rid))

	wb = Workbook()
	OUTFILE = "dominion.xlsx"

	ws1 = wb.active

	style_wrapText = Alignment(wrapText = True)
	style_headFont = Font(bold = True, size = 14)
	style_bodyFont = Font(bold = True, size = 12)
	style_headColor = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')

	ws1.title = "All Recommendations"
	ws1['A1'] = "NAME"
	ws1['B1'] = "DESCRIPTION"
	for rec in recObjs:
		row = (rec.name, rec.__str__())
		ws1.append(row)
	for cell in ws1['1:1']:
		cell.font = style_headFont
		cell.fill = style_headColor

	for cell in ws1['B:B']:
		cell.alignment = style_wrapText

	ws1.column_dimensions['A'].width = float(25)
	ws1.column_dimensions['B'].width = float(100)
	ws1.freeze_panes = 'A2'

	ws2 = wb.create_sheet("All Cards", 0)
	ws2['A1'] = "Name"
	ws2['B1'] = "Description"
	for card in cardObjs:
		row = (card.cardName, card.__str__())
		ws2.append(row)
	for cell in ws2['1:1']:
		cell.font = style_headFont
		cell.fill = style_headColor

	for cell in ws2['B:B']:
		cell.alignment = style_wrapText

	ws2.column_dimensions['A'].width = float(20)
	ws2.column_dimensions['B'].width = float(100)
	ws2.freeze_panes = 'A2'
	
	ws3 = wb.create_sheet("All Sets", 0)
	ws3['A1'] = "Name"
	ws3['B1'] = "Description"
	for cset in setObjs:
		row = (cset.setName, cset.__str__())
		ws3.append(row)
	for cell in ws3['1:1']:
		cell.font = style_headFont
		cell.fill = style_headColor

	for cell in ws3['B:B']:
		cell.alignment = style_wrapText

	ws3.column_dimensions['A'].width = float(15)
	ws3.column_dimensions['B'].width = float(100)
	ws3.freeze_panes = 'A2'

	wb.save(OUTFILE)

def makeFile_card(cid):
	card = Card(cid)
	cset = Set(card.sid)
	rids = []

	sql = """ SELECT rid FROM cardsInRecs WHERE cid = %s """
	cur.execute(sql, (cid,))
	results = cur.fetchall()

	for result in results:
		rids.append(result['rid'])

	wb = Workbook()
	OUTFILE = "card_{}.xlsx".format(card.cardName.lower().strip().replace('\'', ''))

	ws1 = wb.active

	style_wrapText = Alignment(wrapText = True)
	style_headFont = Font(bold = True, size = 14)
	style_bodyFont = Font(bold = True, size = 12)
	style_headColor = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')

	ws1.title = "{}".format(card.cardName.strip().replace('\'', ''))

	ws1['A1'] = "CARD NAME"
	ws1['B1'] = "SET"
	ws1['C1'] = "COST"
	ws1['D1'] = "TYPE"
	ws1['E1'] = "ILLUSTRATOR"
	ws1['F1'] = "DESCRIPTION"

	ws1['A2'] = card.cardName
	ws1['B2'] = cset.setName
	ws1['C2'] = card.cardCost
	ws1['D2'] = card.cardTypes
	ws1['E2'] = card.cardIllustrators
	ws1['F2'] = '{}'.format(card.cardDescription)

	ws1['A4'] = "SET NAME"
	ws1['B4'] = "# CARDS"
	ws1['C4'] = "THEME(S)"
	ws1['D4'] = "RELEASE DATE"
	ws1['E4'] = "BOX ILLUSTRATOR"

	ws1['A5'] = cset.setName
	ws1['B5'] = cset.setCardnumber
	ws1['C5'] = cset.setThemes
	ws1['D5'] = cset.setRelease
	ws1['E5'] = cset.setCoverart

	ws1['A7'] = "REC of 10 NAME"
	ws1['B7'] = "SETS"
	ws1['C7'] = "CARDS"

	for rid in rids:
		rec = Rec(rid)
		cardString = ""
		for i in rec.recCards:
			if i.cardName == rec.recCards[0].cardName:
				cardString += "{}".format(i.cardName)
			else:
				cardString += ", {}".format(i.cardName)
		row = (rec.name, rec.sets, cardString)
		ws1.append(row)

	for cell in ws1['1:1']:
		cell.font = style_headFont
		cell.fill = style_headColor

	for cell in ws1['4:4']:
		cell.font = style_headFont
		cell.fill = style_headColor

	for cell in ws1['7:7']:
		cell.font = style_headFont
		cell.fill = style_headColor

	ws1.column_dimensions['A'].width = float(25)
	ws1.column_dimensions['B'].width = float(25)
	ws1.column_dimensions['C'].width = float(25)
	ws1.column_dimensions['D'].width = float(25)
	ws1.column_dimensions['E'].width = float(25)
	ws1.column_dimensions['F'].width = float(25)

	wb.save(OUTFILE)

def makeFile_set(cid):
	card = Card(cid)
	cset = Set(card.sid)
	makeSet_cards = []

	sql = """ SELECT cid FROM cards WHERE sid = %s """
	cur.execute(sql, (card.sid,))
	results = cur.fetchall()

	for result in results:
		makeSet_cards.append(Card(result['cid']))

	wb = Workbook()
	OUTFILE = "set_{}.xlsx".format(cset.setName.lower().strip().replace('\'', ''))

	ws1 = wb.active

	style_wrapText = Alignment(wrapText = True)
	style_headFont = Font(bold = True, size = 14)
	style_bodyFont = Font(bold = True, size = 12)
	style_headColor = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')

	ws1.title = "{}".format(cset.setName.strip().replace('\'', ''))

	ws1['A1'] = "SET NAME"
	ws1['B1'] = "# CARDS"
	ws1['C1'] = "THEME(S)"
	ws1['D1'] = "RELEASE DATE"
	ws1['E1'] = "BOX ILLUSTRATOR"

	ws1['A2'] = cset.setName
	ws1['B2'] = cset.setCardnumber
	ws1['C2'] = cset.setThemes
	ws1['D2'] = cset.setRelease
	ws1['E2'] = cset.setCoverart

	ws1['A4'] = "CARD NAME"
	ws1['B4'] = "COST"
	ws1['C4'] = "TYPE"
	ws1['D4'] = "ILLUSTRATOR"
	ws1['E4'] = "DESCRIPTION"

	for card in makeSet_cards:
		row = (card.cardName, card.cardCost, card.cardTypes, card.cardIllustrators, card.cardDescription)
		ws1.append(row)

	for cell in ws1['1:1']:
		cell.font = style_headFont
		cell.fill = style_headColor

	for cell in ws1['4:4']:
		cell.font = style_headFont
		cell.fill = style_headColor

	ws1.column_dimensions['A'].width = float(25)
	ws1.column_dimensions['B'].width = float(25)
	ws1.column_dimensions['C'].width = float(25)
	ws1.column_dimensions['D'].width = float(25)
	ws1.column_dimensions['E'].width = float(25)

	wb.save(OUTFILE)

def makeFile_recs(cid):
	card = Card(cid)
	cset = Set(card.sid)
	rids = []

	sql = """ SELECT rid FROM cardsInRecs WHERE cid = %s """
	cur.execute(sql, (cid,))
	results = cur.fetchall()

	for result in results:
		rids.append(result['rid'])

	if len(rids) >= 1:
		noRecs = 1

		wb = Workbook()
		ws1 = wb.active
		ws1.title = "List of sets for {}".format(card.cardName)

		OUTFILE = "recs_{}.xlsx".format(card.cardName.lower().strip().replace('\'', ''))
		style_wrapText = Alignment(wrapText = True)
		style_headFont = Font(bold = True, size = 14)
		style_bodyFont = Font(bold = True, size = 12)
		style_headColor = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')

		ws1['A1'] = "REC of 10 NAME"
		ws1['B1'] = "SETS"

		for cell in ws1['1:1']:
			cell.font = style_headFont
			cell.fill = style_headColor

		ws1.column_dimensions['A'].width = float(25)
		ws1.column_dimensions['B'].width = float(25)

		for rid in rids:
			row = (Rec(rid).name, Rec(rid).sets)
			ws1.append(row)

		for rid in rids:
			recCard_list = []
			
			sql = """ SELECT cid FROM cardsInRecs WHERE rid = %s """
			cur.execute(sql, (rid,))
			results = cur.fetchall()

			for result in results:
				recCard_list.append(Card(result['cid']))

			ws = wb.create_sheet("{}".format(Rec(rid).name.strip().replace('\'', '')))

			ws['A1'] = "CARD NAME"
			ws['B1'] = "SET"
			ws['C1'] = "COST"
			ws['D1'] = "TYPE"
			ws['E1'] = "ILLUSTRATOR"
			ws['F1'] = "DESCRIPTION"

			ws.column_dimensions['A'].width = float(25)
			ws.column_dimensions['B'].width = float(25)
			ws.column_dimensions['C'].width = float(25)
			ws.column_dimensions['D'].width = float(25)
			ws.column_dimensions['E'].width = float(25)
			ws.column_dimensions['F'].width = float(25)

			for cell in ws['1:1']:
				cell.font = style_headFont
				cell.fill = style_headColor

			for card in recCard_list:
				row = (card.cardName, cset.setName, card.cardCost, card.cardTypes, card.cardIllustrators, '{}'.format(card.cardDescription))
				ws.append(row)

			wb.save(OUTFILE)

	else:
		noRecs = 0

	return(noRecs)	

## /Data Fetching ##
# /FUNCTIONS #

class Set(object):
	def __init__(self, sid):
		self.sid = sid

		cur.execute(""" SELECT * FROM csets WHERE sid = %s """, (sid,))
		result = cur.fetchall()
		self.setName = result[0]['name'].strip()
		self.setCardnumber = result[0]['cardnumber'].strip()
		self.setThemes = result[0]['themes'].strip().replace('\n', ', ').replace(' ,', '').replace('  ', ' ').replace(',  ','')
		self.setRelease = result[0]['release'].strip()
		self.setCoverart = result[0]['coverart'].strip()
		self.setCardset = []

	def buildSet(self):
		cur.execute(""" SELECT cid FROM cards WHERE sid = %s """, (self.sid,))
		result = cur.fetchall()
		self.setCards = []
		cidList = []
		for i in result:
			cidList.append(i['cid'])
		for cid in cidList:
			self.setCards.append(Card(cid))

		return(self.setCards)

	def __contains__(self, inSetName):
		self.setCardset = self.buildSet()

		for card in self.setCardset:
			if inSetName == card.cardName:
				print("The card {} is in the set {}.".format(inSetName, self.setName))
				return(True)
			else:
				print("The card {} is NOT in the set {}.".format(inSetName, self.setName))
				return(False)


	def __str__(self):
		return("{} is a {}-card Dominion set with the theme(s): {}. It was released {}, with cover art by {}.".format(self.setName, self.setCardnumber, self.setThemes, self.setRelease, self.setCoverart))

class Card(object):
	def __init__(self, cid):
		self.cid = cid

		cur.execute(""" SELECT * FROM cards WHERE cid = %s """, (cid,))
		result = cur.fetchall()
		self.sid = result[0]['sid']
		self.cardName = result[0]['name'].strip()
		self.cardCost = result[0]['cost'].strip()
		self.cardTypes = result[0]['types'].strip()
		self.cardIllustrators = result[0]['illustrators'].strip()
		self.cardDescription = result[0]['description'].strip()
		self.cardRecommendations = []

		cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (self.sid,))
		result = cur.fetchall()
		self.setName = result[0]['name'].strip()

	def buildRecsList(self):
		cur.execute(""" SELECT rid FROM cardsInRecs WHERE cid = %s """, (cid,))
		result = cur.fetchall()

		ridList = []
		for i in result:
			ridList.append(i['rd'])
		for rid in ridList:
			self.cardRecommendations.append(Rec(rid))

		return(self.cardRecommendations)

	def __str__(self):
		return("'{}' is a card from the {} set, illustrated by {}.\nIt is a card of type(s): {}. It costs {}.\nDescription:\n'{}'".format(self.cardName, self.setName, self.cardIllustrators, self.cardTypes, self.cardCost, self.cardDescription))

	def __repr__(self):
		return("Card: {}".format(self.cardName))

class Rec(object):
	def __init__(self, rid):
		self.rid = rid

		cur.execute(""" SELECT * FROM recommendations WHERE rid = %s """, (rid,))
		result = cur.fetchall()
		self.name = result[0]['name']
		self.sets = ""
		sets = [str(result[0]['set1'])]
		if result[0]['set2']:
			sets.append(str(result[0]['set2']))
		if result[0]['set3']:
			sets.append(str(result[0]['set3']))

		if len(sets) == 1:
			cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (sets[0],))
			result = cur.fetchall()
			self.sets += result[0]['name']

		elif len(sets) == 2:
			cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (sets[0],))
			result = cur.fetchall()
			self.sets += result[0]['name']

			cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (sets[1],))
			result = cur.fetchall()
			self.sets += " & {}".format(result[0]['name'])

		elif len(sets) == 3:
			cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (sets[0],))
			result = cur.fetchall()
			self.sets += result[0]['name']

			cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (sets[1],))
			result = cur.fetchall()
			self.sets += ", {}".format(result[0]['name'])

			cur.execute(""" SELECT name FROM csets WHERE sid = %s """, (sets[2],))
			result = cur.fetchall()
			self.sets += ", & {}".format(result[0]['name'])

		cur.execute(""" SELECT cid FROM cardsInRecs WHERE rid = %s """, (rid,))
		result = cur.fetchall()
		self.recCards = []
		cidList = []
		for i in result:
			cidList.append(i['cid'])
		for cid in cidList:
			self.recCards.append(Card(cid))

	def __str__(self):
		cardNames = ""
		num = len(self.recCards)
		num = num - 1
		for card in self.recCards:
			if card != self.recCards[num]:
				cardNames += "{}, ".format(card.cardName)
			else:
				cardNames += "{}".format(card.cardName)

		return("{} is a recommended Kingdom for the set(s) {}. It includes the cards: {}.".format(self.name, self.sets, cardNames))

# EXECUTE
## Get connection
conn, cur = getConnection_andCursor()

## Make database tables
if GETNEWDATA:
	makeTables()

# # Make list of links from which to collect data
if GETNEWDATA:
	setLinks = getSetLinks()
	cardLinks = getCardLinks()

# Make list of soup objects from list of links
if GETNEWDATA:
	setData = getSetData(setLinks)
	cardData = getCardData(cardLinks)
	recData = getRecData(setLinks)

# Insert data into database
if GETNEWDATA:
	for cset in setData:
		insertSet(cset)

	for card in cardData:
		insertCard(card)

	for rec in recData:
		insertRec(rec)

# Make lists of ids
allCards = fetchAll_cardIDs()
allSets = fetchAll_setIDs()
allRecs = fetchAll_recIDs()


print(type(Card(1)))

# Prompt #
promptMsg = originalPrompt

while not DONE:
	userInput = input(promptMsg)

	if userInput == 'help':
		promptMsg = "\nYou can type:\n'list': generate a csv file with a list of all Dominion cards, sets, and recommended sets of 10\n'random': start viewing information about a random card\n'exit': leave the program\n"

	elif userInput.lower() == 'list':
		makeList()
		print("\nPlease look for the file 'dominion.xlsx' in the same directory as dominionExplorer.py.")

	elif userInput.lower() == 'random':
		localMsg = "You can type:\n'card': print all info about this card to an excel file\n'recs': print all the recommmended sets of 10 this card is in to an excel file\n'set': print all info about the set this card belongs to, including a list of all other cards in the set\n'back': go back to main menu\n"
		rn = random.randrange(0, (len(allCards) - 1))
		randomCard = Card(rn)
		
		print("\n{}".format(randomCard))
		promptMsg = "\nCURRENT CARD: {}. {}".format(randomCard.cardName, localMsg)

		localExit = False
		
		while not localExit:
			userInput = input(promptMsg)

			if userInput.lower() == 'card':
				makeFile_card(randomCard.cid)
				print("\nPlease look for the file 'card_{}.xlsx' in the same directory as dominionExplorer.py.".format(randomCard.cardName.lower().strip().replace('\'', '')))

			elif userInput.lower() == 'recs':
				if makeFile_recs(randomCard.cid) == False:
					print("\nSorry, there are no recommendations for '{}'!".format(randomCard.cardName.lower().strip().replace('\'', '')))
				else:
					print("\nPlease look for the 'file recs_{}.xlsx' in the same directory as dominionExplorer.py.".format((randomCard.cardName.lower().strip().replace('\'', ''))))

			elif userInput.lower() == 'set':
				makeFile_set(randomCard.cid)
				print("\nPlease look for the file 'set_{}.xlsx' in the same directory as dominionExplorer.py.".format(randomCard.setName.lower().strip().replace('\'', '')))

			elif userInput.lower() == 'back':
				localExit = True

			else:
				print("\nSorry, I didn't recognize command '{}'.".format(userInput))

		promptMsg = originalPrompt

	elif userInput.lower() == "done":
		DONE = True

	else:
		localMsg = "You can type:\n'card': print all info about this card to an excel file\n'recs': print all the recommmended sets of 10 this card is in to an excel file\n'set': print all info about the set this card belongs to, including a list of all other cards in the set\n'back': go back to main menu\n"

		sql = """ SELECT cid FROM cards WHERE name = %s """
		cur.execute(sql, (userInput,))
		result = cur.fetchall()

		if result:		

			userCard = Card(result[0]['cid'])
			
			print("\n{}".format(userCard))
			promptMsg = "\nCURRENT CARD: {}. {}".format(userCard.cardName, localMsg)

			localExit = False
			
			while not localExit:
				userInput = input(promptMsg)

				if userInput.lower() == 'card':
					makeFile_card(userCard.cid)
					print("\nPlease look for the file 'card_{}.xlsx' in the same directory as dominionExplorer.py.".format(userCard.cardName.lower().strip().replace('\'', '')))

				elif userInput.lower() == 'recs':
					if makeFile_recs(userCard.cid) == False:
						print("\nSorry, there are no recommendations for {}!".format(userCard.cardName))
					else:
						makeFile_recs(randCard.cid)
						print("\nPlease look for the file 'recs_{}.xlsx' in the same directory as dominionExplorer.py.".format((userCard.cardName.lower().strip().replace('\'', ''))))

				elif userInput.lower() == 'set':
					makeFile_set(userCard.cid)
					print("\nPlease look for the file 'set_{}.xlsx' in the same directory as dominionExplorer.py.".format(userCard.setName.lower().strip().replace('\'', '')))

				elif userInput.lower() == 'back':
					localExit = True

				else:
					print("\nSorry, I didn't recognize command '{}'.".format(userInput))

			promptMsg = originalPrompt
			userInput.lower()
		else:
			print("\nSorry, could not find a card called '{}'. Please check case (search terms for cards are case-sensitive).".format(userInput))
